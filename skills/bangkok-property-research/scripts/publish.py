#!/usr/bin/env python3
"""
Publish Bangkok property rows to (a) a local .xlsx artifact with embedded picture
thumbnails, and (b) — if BANGKOK_SHEET_WEBHOOK is set — a live Google Sheet via an
Apps Script Web App, which keeps ONE persistent spreadsheet updated each week.

Usage:
    python3 publish.py --rows rows.json --xlsx out.xlsx

rows.json = a JSON array of unit objects using the SKILL.md schema. Missing fields
are tolerated (rendered blank). This script does NOT invent data.
"""
import argparse, io, json, os, sys, datetime, urllib.parse

def log(*a): print("[publish]", *a, file=sys.stderr)

# ---- column order shown in the sheet ----
COLUMNS = [
    ("score", "Score"), ("title", "Title"), ("district", "District"),
    ("position", "Position"), ("price_usd", "Price USD"), ("price_thb", "Price THB"),
    ("sqm", "sqm"), ("price_per_sqm_thb", "THB/sqm"), ("bedrooms", "Beds"),
    ("foreign_freehold", "Foreign Freehold"), ("link", "Link"),
    ("image_url", "Picture"), ("source", "Source"), ("notes", "Notes"),
    ("first_seen", "First seen"), ("last_seen", "Last seen"),
]

def today():
    return datetime.date.today().isoformat()

def clean(rows):
    out = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        r.setdefault("first_seen", today())
        r.setdefault("last_seen", today())
        out.append(r)
    # highest score first; None scores sink to the bottom
    out.sort(key=lambda r: (r.get("score") is not None, r.get("score") or 0), reverse=True)
    return out

# ---------------------------------------------------------------- xlsx
def write_xlsx(rows, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    thumbs_ok = _try_import_thumbs()

    def build(ws, subset, title):
        ws.title = title
        headers = [h for _, h in COLUMNS]
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(vertical="center")
        pic_col = [k for k, _ in COLUMNS].index("image_url") + 1
        link_col = [k for k, _ in COLUMNS].index("link") + 1
        for r in subset:
            values = []
            for key, _ in COLUMNS:
                v = r.get(key)
                values.append("" if v is None else v)
            ws.append(values)
            row_idx = ws.max_row
            ws.row_dimensions[row_idx].height = 60
            # hyperlink the Link cell
            link = r.get("link")
            if link:
                lc = ws.cell(row=row_idx, column=link_col)
                lc.hyperlink = link
                lc.value = "open"
                lc.font = Font(color="0563C1", underline="single")
            # embed a thumbnail if we could fetch it, else leave the URL as text
            img_url = r.get("image_url")
            placed = False
            if img_url and thumbs_ok:
                placed = _embed_thumb(ws, row_idx, pic_col, img_url)
            if not placed:
                ws.cell(row=row_idx, column=pic_col).value = img_url or ""
        # column widths
        widths = {"Title": 26, "District": 22, "Position": 26, "Notes": 40,
                  "Picture": 16, "Link": 8, "Foreign Freehold": 18}
        for i, (_, h) in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 12)
        ws.freeze_panes = "A2"

    build(wb.active, rows, "Master")
    week = [r for r in rows if r.get("last_seen") == today()]
    build(wb.create_sheet(), week, "This Week")

    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)
    log(f"xlsx written: {path}  ({len(rows)} master rows, {len(week)} this week)")

_THUMB_CACHE = {}
def _try_import_thumbs():
    try:
        import requests, PIL  # noqa
        return True
    except Exception:
        log("Pillow/requests missing — thumbnails will be URLs, not images")
        return False

def _embed_thumb(ws, row_idx, col_idx, url):
    try:
        import requests
        from PIL import Image as PILImage
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
        if url not in _THUMB_CACHE:
            # Portal CDNs hotlink-protect: cdn.fazwaz.com 403s any request without a
            # Referer from fazwaz.com. Send one derived from the image host.
            from urllib.parse import urlsplit
            _host = urlsplit(url).netloc
            _ref = "https://www.fazwaz.com/" if "fazwaz" in _host else f"https://{_host}/"
            resp = requests.get(url, timeout=12, headers={
                "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                               "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124 Safari/537.36"),
                "Accept": "image/avif,image/webp,image/*,*/*;q=0.8",
                "Referer": _ref,
                "Sec-Fetch-Dest": "image",
                "Sec-Fetch-Mode": "no-cors",
                "Sec-Fetch-Site": "cross-site",
            })
            resp.raise_for_status()
            im = PILImage.open(io.BytesIO(resp.content)).convert("RGB")
            im.thumbnail((110, 78))
            buf = io.BytesIO(); im.save(buf, format="PNG"); buf.seek(0)
            _THUMB_CACHE[url] = buf.getvalue()
        bio = io.BytesIO(_THUMB_CACHE[url])
        xi = XLImage(bio)
        anchor = f"{get_column_letter(col_idx)}{row_idx}"
        ws.add_image(xi, anchor)
        return True
    except Exception as e:
        log(f"thumb fail {url[:60]}: {e}")
        return False

# ---------------------------------------------------------------- webhook
def push_webhook(rows, url):
    import requests
    payload = {"generated": today(), "columns": COLUMNS, "rows": rows}
    try:
        resp = requests.post(url, json=payload, timeout=60,
                             headers={"Content-Type": "application/json"})
        resp.raise_for_status()
        body = resp.text.strip()
        log("webhook ok:", body[:200])
        # Apps Script returns the spreadsheet URL as text or JSON {url:...}
        try:
            j = json.loads(body)
            return j.get("url") or body
        except Exception:
            return body
    except Exception as e:
        log("webhook FAILED:", e)
        return None

# ---------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--rows", required=True)
    ap.add_argument("--xlsx", required=True)
    args = ap.parse_args()

    with open(args.rows) as f:
        rows = clean(json.load(f))

    write_xlsx(rows, args.xlsx)

    hook = os.environ.get("BANGKOK_SHEET_WEBHOOK", "").strip()
    if hook:
        sheet_url = push_webhook(rows, hook)
        if sheet_url:
            print(f"LIVE_SHEET_URL={sheet_url}")
    else:
        log("BANGKOK_SHEET_WEBHOOK not set — skipped live Google Sheet. "
            "xlsx artifact is the deliverable (upload to Drive via the Drive MCP if desired).")
    print(f"XLSX_PATH={os.path.abspath(args.xlsx)}")

if __name__ == "__main__":
    main()
